import openpyxl
import datetime

# show how to read from an existing workbook
wb_read = openpyxl.load_workbook('example.xlsx')
sheet = wb_read['Sheet1']

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

# show how to create a new workbook
# create a workbook
wb_write = openpyxl.Workbook()

# get the reference of an active sheet
sheet_write = wb_write.active

# set the title of the active sheet
sheet_write.title = 'Spam Spam Spam'

sheet_write['A1'] = 'Hello world!'

# create the workbook
wb_write.save('example_copy.xlsx')

# show how to modify an existing workbook
# Open an xlsx for reading
wb = openpyxl.load_workbook('example_modify.xlsx')
# Get the current Active Sheet
# ws = wb.active
wb.active.cell(row=6,column=7).value = 'Harry'
wb.active.cell(row=8,column=8).value = 'Lo'

# write the timestamp
wb.active.cell(row=1,column=1).value = datetime.datetime.now()

# save the csb file
wb.save('example_modify.xlsx')